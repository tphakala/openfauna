#!/usr/bin/env python3
"""Replace English-fallback common names with real localized names.

OpenFauna was bootstrapped from the BirdNET taxonomy, which fills every locale it
cannot translate with the English common name. The result is thousands of
"English echoes": entries in a non-English locale whose value is verbatim the
English name (e.g. fr.json "Pipistrellus tenuis" -> "least pipistrelle"). The
earlier shadow backfill (scripts/backfill_vernacular.py) only touched entries
equal to the scientific name, so it never saw these. This script targets them.

Source cascade per (species, locale):
  1. IOC World Bird List (birds only; CC BY 3.0). Authoritative, correctly
     accented, one curated name per language. Trusted: applied without QA.
  2. GBIF Backbone vernacular names (ISO 639-3; CC0 / source license).
  3. Wikidata taxon common name P1843 (CC0), alias-aware.
  GBIF/Wikidata are crowd-sourced and carry fabricated names (a translated genus
  word plus a latinized epithet, e.g. "Pipistrelle abramusi"), misspellings and
  wrong-language entries, so their candidates pass extra filters and are tagged
  source!=IOC for a separate quality pass before they are trusted.

It only rewrites an English-echo or a scientific-name shadow; it never overwrites
a curated translation. Genuinely unnameable species (obscure exotics no source
names in the target language) are left as the English fallback on purpose: they
are not native to that locale, which is the bar set for this dataset.

Usage:
  python3 scripts/backfill_native.py                 # dry run + report
  python3 scripts/backfill_native.py --report out.md
  python3 scripts/backfill_native.py --apply          # write locale files
  python3 scripts/backfill_native.py --apply --sources ioc   # IOC only
"""

import argparse
import json
import os
import re
import sys
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from backfill_v24_open import gbif_vernacular, gbif_pick, normalize  # noqa: E402
from backfill_vernacular import (  # noqa: E402
    LOCALE_LANG, CASE_STYLE, clean_name, is_rank_label, is_sci_echo,
    wikidata_batch, wikidata_aliased, wd_pick, _ascii_fold, _lev,
)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LOCALE_DIR = os.path.join(ROOT, "data", "locales")
META_FILE = os.path.join(ROOT, "data", "metadata.json")
ALIASES_FILE = os.path.join(ROOT, "data", "aliases.json")
CURATED_FILE = os.path.join(ROOT, "data", "curated_names.json")
REJECTS_FILE = os.path.join(ROOT, "data", "validation", "gbif_rejects.json")
GBIF_CACHE = os.path.join(os.path.expanduser("~"), ".cache", "openfauna", "gbif_vern_cache.json")
WD_CACHE = os.path.join(os.path.expanduser("~"), ".cache", "openfauna", "wikidata_vern_cache.json")
IOC_XLSX = os.environ.get("OPENFAUNA_IOC", "/tmp/openfauna-work/ioc.xlsx")

# OpenFauna locale -> IOC column header. Regional variants share the base column
# (a generic Spanish/Portuguese IOC name beats an English fallback). Locales with
# no IOC column (cy, hi_in, vi_vn) fall through to GBIF/Wikidata.
IOC_COL = {
    "ca": "Catalan", "zh_cn": "Chinese", "hr": "Croatian", "cs": "Czech",
    "da": "Danish", "nl": "Dutch", "fi": "Finnish", "fr": "French",
    "de": "German", "it": "Italian", "ja": "Japanese", "lt": "Lithuanian",
    "no": "Norwegian", "pl": "Polish", "pt": "Portuguese (Lusophone)",
    "pt_pt": "Portuguese (Portuguese)", "ru": "Russian", "sr": "Serbian",
    "sk": "Slovak", "es": "Spanish", "es_es": "Spanish", "es_ec": "Spanish",
    "es_mx": "Spanish", "sv": "Swedish", "tr": "Turkish", "uk": "Ukrainian",
    "af": "Afrikaans", "ar": "Arabic", "bg": "Bulgarian", "et": "Estonian",
    "et_ee": "Estonian", "el": "Greek", "he": "Hebrew", "hu": "Hungarian",
    "is": "Icelandic", "id": "Indonesian", "ko": "Korean", "lv_lv": "Latvian",
    "ml": "Malayalam", "fa": "Persian", "ro": "Romanian", "sl": "Slovenian",
    "th": "Thai",
}

BINOMIAL = re.compile(r"^[A-Z][a-z]+ [a-z]+")
# Latin grammatical endings a fabricator tacks onto an epithet to "localize" it.
_LAT_SUFFIX = re.compile(r"(i|ii|us|a|ae|us|is|ensis|orum|arum|anus|inus)$")

# Locales whose species common names are written all-lowercase by convention.
LOWER_LOCALES = ("fi", "sv", "no", "cs")


def clean_case(name, loc, trusted):
    """Normalize a candidate's punctuation and capitalization for storage.

    Always replaces em/en dashes with a plain hyphen (the dataset forbids them)
    and collapses whitespace. Trusted names (IOC, curated, reviewer-fixed) are
    already correctly cased for their language - including capitalized eponyms
    like 'Tinamu de Darwin' - so their casing is preserved. Untrusted GBIF/Wikidata
    names get minimal normalization: lowercased for the all-lowercase locales,
    otherwise just a capitalized first letter with interior casing kept (so a
    proper noun is never folded to lowercase, the failure mode of sentence-casing).
    """
    name = name.replace("—", "-").replace("–", "-").replace("‑", "-")
    name = " ".join(name.split())
    if not name:
        return name
    if trusted:
        return name
    if loc in LOWER_LOCALES:
        return name.lower()
    return name[0].upper() + name[1:]


def load_json(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def save_json(path, obj):
    with open(path, "w", encoding="utf-8") as f:
        f.write(json.dumps(obj, indent=2, ensure_ascii=False, sort_keys=True) + "\n")


def locale_files():
    return {
        f[:-5]: load_json(os.path.join(LOCALE_DIR, f))
        for f in os.listdir(LOCALE_DIR) if f.endswith(".json")
    }


# ---- IOC ----

def load_ioc():
    import openpyxl
    wb = openpyxl.load_workbook(IOC_XLSX, read_only=True, data_only=True)
    try:
        ws = wb["List"] if "List" in wb.sheetnames else wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        hdr = list(next(it))
        sci_col = hdr.index("IOC_15.2")
        col_idx = {h: i for i, h in enumerate(hdr) if h}
        rows = {}
        for row in it:
            # openpyxl read-only mode truncates a row at its last non-empty cell, so a
            # short row may not reach sci_col; guard the index.
            if len(row) > sci_col and row[sci_col]:
                rows[str(row[sci_col]).strip()] = row
        return rows, col_idx
    finally:
        wb.close()


def ioc_name(ioc_rows, col_idx, aliases, sci, loc):
    col = IOC_COL.get(loc)
    if not col:
        return None
    ci = col_idx.get(col)
    if ci is None:
        return None
    for key in (sci, aliases.get(sci)):
        if key and key in ioc_rows:
            row = ioc_rows[key]
            v = row[ci] if len(row) > ci else None  # short rows are truncated by openpyxl
            if v and str(v).strip():
                s = str(v).strip()
                # IOC leaves the scientific name in the cell when a language has no
                # common name for the species; that is a shadow, not a translation.
                if s.lower() == key.lower() or s.lower() == sci.lower():
                    return None
                return s
    return None


# ---- fabrication filter for GBIF/Wikidata ----

def is_fabricated(name, sci):
    """True when a vernacular candidate is a latinized echo of the epithet: a
    translated genus word followed by the scientific epithet with a Latin ending,
    e.g. 'Pipistrelle abramusi' for Pipistrellus abramus. is_sci_echo only checks
    whole-string edit distance and misses these because the genus word differs."""
    epithet = sci.split(" ")[-1].lower()
    stem = _LAT_SUFFIX.sub("", epithet)
    if len(stem) < 4:
        return False
    for tok in _ascii_fold(name.lower()).split():
        tstem = _LAT_SUFFIX.sub("", re.sub(r"[^a-z]", "", tok))
        if len(tstem) < 4:
            continue
        if _lev(tstem, stem) <= 1:
            return True
    return False


# ---- targets ----

def is_english_echo(value, en_real_val, key):
    if en_real_val is None or not isinstance(value, str):
        return False
    v = value.strip()
    return v == en_real_val.strip() and v.lower() != key.lower()


_LATIN_VALUE = re.compile(r"^[A-Z][a-z]+ [a-z]+( [a-z]+)?$")


def is_sci_value(value, key, genera):
    """True when the value is a scientific name in disguise, not a common name: a
    Latin binomial/trinomial synonym of the key. It must start with a known genus
    (so a real vernacular name ending in a Latin-looking word, e.g. Afrikaans
    'Amerikaanse bison', is not mistaken for one) and then either share the key's
    epithet (a reclassification, e.g. 'Bufo boreas' for Anaxyrus boreas) or keep the
    same genus with a gender-variant epithet (e.g. 'Adenomera nanus' for nana)."""
    v = value.strip()
    if not _LATIN_VALUE.match(v) or v.lower() == key.lower():
        return False
    vt, kt = v.split(" "), key.split(" ")
    if vt[0] not in genera:
        return False
    if vt[-1] == kt[-1]:
        return True
    return vt[0] == kt[0] and _lev(vt[1], kt[1]) <= 2


def build_targets(data, en_real, genera):
    """{species: {locale: kind}} for English-echo and sci-name shadow entries."""
    targets = {}
    for loc, d in data.items():
        if loc.startswith("en"):
            continue
        for k, v in d.items():
            if not isinstance(v, str) or not BINOMIAL.match(k):
                continue
            vs = v.strip()
            if vs.lower() == k.lower() or is_sci_value(vs, k, genera):
                kind = "shadow"
            elif is_english_echo(v, en_real.get(k), k):
                kind = "english"
            else:
                continue
            targets.setdefault(k, {})[loc] = kind
    return targets


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--report", help="write a markdown change report")
    ap.add_argument("--sources", default="ioc,gbif,wikidata",
                    help="comma list of sources to use, in priority order")
    ap.add_argument("--workers", type=int, default=8)
    ap.add_argument("--no-net", action="store_true", help="use only cached GBIF/Wikidata")
    args = ap.parse_args()
    use = set(args.sources.split(","))

    data = locale_files()
    en = data["en"]
    en_real = {k: v for k, v in en.items()
               if isinstance(v, str) and v.strip() and v.strip().lower() != k.lower()}
    meta = load_json(META_FILE)
    aliases = load_json(ALIASES_FILE)
    curated = load_json(CURATED_FILE).get("names", {}) if os.path.exists(CURATED_FILE) else {}
    rejects_doc = load_json(REJECTS_FILE) if os.path.exists(REJECTS_FILE) else {}
    qa_reject = {lang: set(v) for lang, v in rejects_doc.get("reject", {}).items()}
    qa_fix = rejects_doc.get("fix", {})
    genera = {k.split(" ")[0] for k in en if BINOMIAL.match(k)}
    targets = build_targets(data, en_real, genera)
    # A curated name is authoritative: it overrides any slot for that species that is
    # not already a genuine translation - an absent key, a scientific-name shadow or
    # synonym, or an English fallback (matched case-insensitively, so a lowercased
    # English placeholder is caught too). It never replaces a real pre-existing
    # localized name. This case-insensitive match is scoped to curated species only,
    # so it does not balloon the global (exact-match) English-echo target set.
    for sci, locs in curated.items():
        ev = en_real.get(sci)
        for loc in locs:
            if loc not in data:
                continue
            cur = data[loc].get(sci)
            genuine = (isinstance(cur, str) and cur.strip()
                       and cur.strip().lower() != sci.lower()
                       and not (ev and cur.strip().lower() == ev.strip().lower())
                       and not is_sci_value(cur.strip(), sci, genera))
            # Still override a "genuine" value when the curated name is the same name
            # with corrected accents/casing (e.g. "Minioptere" -> "Minioptère"), but
            # never when it is a different name (so a valid synonym is left alone).
            if genuine and isinstance(cur, str):
                accent_fix = _ascii_fold(cur.strip().lower()) == _ascii_fold(locs[loc].strip().lower())
            else:
                accent_fix = False
            if not genuine or accent_fix:
                targets.setdefault(sci, {})[loc] = "curated"
    species = sorted(targets)
    npairs = sum(len(v) for v in targets.values())
    print(f"targets: {len(species)} species, {npairs} (species, locale) pairs")

    ioc_rows, col_idx = ({}, {})
    if "ioc" in use:
        if not os.path.exists(IOC_XLSX):
            print(f"WARNING: IOC workbook not found at {IOC_XLSX}; skipping the IOC source "
                  "(set OPENFAUNA_IOC to its path to enable authoritative bird names).",
                  file=sys.stderr)
        else:
            ioc_rows, col_idx = load_ioc()
            print(f"IOC rows: {len(ioc_rows)}")

    gbif_cache = load_json(GBIF_CACHE) if os.path.exists(GBIF_CACHE) else {}
    wd_cache = load_json(WD_CACHE) if os.path.exists(WD_CACHE) else {}

    # Only fetch GBIF/Wikidata for species not resolved by IOC in all their slots,
    # to keep the network footprint bounded.
    need_net = []
    for sci in species:
        if not ({"gbif", "wikidata"} & use):
            break
        unresolved = [loc for loc in targets[sci]
                      if not (("ioc" in use) and ioc_name(ioc_rows, col_idx, aliases, sci, loc))]
        if unresolved:
            need_net.append(sci)

    if "gbif" in use and not args.no_net:
        todo = [s for s in need_net if s not in gbif_cache]
        print(f"GBIF lookups: {len(todo)} (cache {len(gbif_cache)})")
        os.makedirs(os.path.dirname(GBIF_CACHE), exist_ok=True)
        # Persist periodically and on any exit so an interruption mid-scrape does not
        # discard hours of fetched vernacular data.
        try:
            with ThreadPoolExecutor(max_workers=args.workers) as ex:
                for i, (sci, out, ok) in enumerate(ex.map(gbif_vernacular, todo)):
                    if ok:
                        gbif_cache[sci] = out
                    if i and i % 500 == 0:
                        print(f"  gbif {i}/{len(todo)}")
                        save_json(GBIF_CACHE, gbif_cache)
        finally:
            save_json(GBIF_CACHE, gbif_cache)

    if "wikidata" in use and not args.no_net:
        wd_todo = [s for s in need_net if s not in wd_cache]
        print(f"Wikidata batch lookups: {len(wd_todo)} (cache {len(wd_cache)})")
        batch = wikidata_batch(wd_todo)
        for sci in wd_todo:
            wd_cache[sci] = batch.get(sci, {})
        os.makedirs(os.path.dirname(WD_CACHE), exist_ok=True)
        save_json(WD_CACHE, wd_cache)

    # Resolve.
    from collections import Counter
    # Names a reviewer flagged as group/family labels are rejected wholesale for the
    # whole language, not just the one species GBIF happened to attach them to.
    qa_reject_names = {lang: {n.lower() for n in v}
                       for lang, v in rejects_doc.get("reject_names", {}).items()}

    # Honor the order the user gave in --sources (a set loses it) for the GBIF vs
    # Wikidata preference; the default "gbif,wikidata" keeps GBIF first.
    net_order = [s for s in args.sources.split(",") if s in ("gbif", "wikidata")]

    def raw_pick(sci, loc):
        """Best non-trusted (GBIF/Wikidata) candidate for a slot, before filters,
        consulting the sources in the order the user requested."""
        iso3, wdl = LOCALE_LANG.get(loc, (None, None))
        for src in net_order:
            if src == "gbif" and iso3:
                p = gbif_pick(gbif_cache.get(sci, {}).get(iso3, []), sci, loc)
                if p:
                    return p, "GBIF"
            elif src == "wikidata" and wdl:
                p = wd_pick(wd_cache.get(sci, {}).get(wdl, []), sci, loc)
                if p:
                    return p, "Wikidata"
        return None, None

    # Pass 1: count how many distinct species each raw GBIF/Wikidata name covers per
    # locale. Counted on the unfiltered candidates so that removing one instance (by a
    # per-species QA reject) cannot mask the rest of a reused group label.
    name_uses = Counter()
    for sci in species:
        for loc in targets[sci]:
            p, src = raw_pick(sci, loc)
            if p:
                name_uses[(loc, clean_case(p, loc, False).lower())] += 1

    # Pass 2: resolve with the cascade and all quality filters.
    changes = []  # (sci, loc, kind, new, source)
    skipped_fab, skipped_group = [], []
    for sci in species:
        for loc, kind in sorted(targets[sci].items()):
            iso3, wdl = LOCALE_LANG.get(loc, (None, None))
            name = source = None
            cn = curated.get(sci, {}).get(loc)
            if cn:
                name, source = cn, "curated"
            if not name and "ioc" in use:
                n = ioc_name(ioc_rows, col_idx, aliases, sci, loc)
                if n:
                    name, source = n, "IOC"
            if not name:
                name, source = raw_pick(sci, loc)
            if not name:
                continue
            fixed = False
            if source not in ("IOC", "curated"):
                if is_rank_label(name) or is_sci_echo(name, sci):
                    continue
                if is_fabricated(name, sci):
                    skipped_fab.append((sci, loc, name, source))
                    continue
                # Per-language QA verdicts (data/validation/gbif_rejects.json).
                if wdl and sci in qa_reject.get(wdl, set()):
                    continue
                if wdl and name.lower() in qa_reject_names.get(wdl, set()):
                    continue
                if wdl and sci in qa_fix.get(wdl, {}):
                    name = qa_fix[wdl][sci]
                    fixed = True
            cleaned = clean_case(name, loc, trusted=source in ("IOC", "curated") or fixed)
            # Safety net: never write the scientific name as a common name (any source).
            if cleaned.strip().lower() == sci.lower():
                continue
            # Group-name leakage: a genuine vernacular name is species-specific, so a
            # GBIF/Wikidata candidate reused for two or more species in the same locale
            # is almost always a family/group label (e.g. fi "australiankehrääjät" =
            # owlet-nightjars). IOC/curated are one-name-per-species, so are exempt.
            if source not in ("IOC", "curated") and name_uses[(loc, cleaned.lower())] >= 2:
                skipped_group.append((sci, loc, kind, cleaned, source))
                continue
            # No-op guard: skip if cleaning leaves it equal to the current value or to
            # the English name (GBIF sometimes returns the English name verbatim).
            cur = data[loc].get(sci)
            if cur is not None and cleaned.strip() == cur.strip():
                continue
            if source not in ("IOC", "curated") and cleaned.strip().lower() == (en_real.get(sci) or "").strip().lower():
                continue
            changes.append((sci, loc, kind, cleaned, source))

    changes.sort()
    by_src = {}
    for _, _, _, _, s in changes:
        by_src[s] = by_src.get(s, 0) + 1
    print(f"\nresolved {len(changes)} of {npairs} pairs: {by_src}")
    print(f"rejected as fabricated: {len(skipped_fab)}; as group/family label: {len(skipped_group)}")

    # Emit the non-IOC changes as JSON so a quality pass (subagents) can review them.
    qa = [{"locale": loc, "sci": sci, "name": name, "source": source}
          for sci, loc, kind, name, source in changes if source not in ("IOC", "curated")]
    qa_path = "/tmp/openfauna-work/gbif_changes.json"
    os.makedirs(os.path.dirname(qa_path), exist_ok=True)
    with open(qa_path, "w", encoding="utf-8") as f:
        json.dump(qa, f, ensure_ascii=False, indent=0)

    if args.report:
        write_report(args.report, species, npairs, changes, skipped_fab, targets, data, en_real, meta)

    if not args.apply:
        print("\nDRY RUN (use --apply to write).")
        return

    touched = set()
    for sci, loc, kind, name, source in changes:
        data[loc][sci] = name
        touched.add(loc)
    for loc in sorted(touched):
        save_json(os.path.join(LOCALE_DIR, f"{loc}.json"), data[loc])
    print(f"\nWROTE {len(touched)} locale files.")


def write_report(path, species, npairs, changes, skipped_fab, targets, data, en_real, meta):
    from collections import Counter
    per_loc = Counter()
    per_loc_src = {}
    for sci, loc, kind, name, source in changes:
        per_loc[loc] += 1
        per_loc_src.setdefault(loc, Counter())[source] += 1
    lines = ["# Native-name backfill report", "",
             f"- target pairs (English-echo + shadow): {npairs}",
             f"- resolved: {len(changes)}",
             f"- rejected as fabricated: {len(skipped_fab)}", "",
             "## Per-locale fills", "",
             "| locale | filled | IOC | GBIF | Wikidata |",
             "| --- | --- | --- | --- | --- |"]
    for loc in sorted(per_loc, key=lambda x: -per_loc[x]):
        s = per_loc_src[loc]
        lines.append(f"| {loc} | {per_loc[loc]} | {s.get('IOC',0)} | {s.get('GBIF',0)} | {s.get('Wikidata',0)} |")
    lines += ["", "## All GBIF/Wikidata names applied (for quality review)", "",
              "| locale | scientific name | filled name | source |",
              "| --- | --- | --- | --- |"]
    for sci, loc, kind, name, source in changes:
        if source not in ("IOC", "curated"):
            lines.append(f"| {loc} | {sci} | {name} | {source} |")
    lines += ["", "## Rejected as fabricated", "",
              "| locale | scientific name | rejected candidate | source |",
              "| --- | --- | --- | --- |"]
    for sci, loc, name, source in skipped_fab:
        lines.append(f"| {loc} | {sci} | {name} | {source} |")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print(f"wrote report to {path}")


if __name__ == "__main__":
    main()
